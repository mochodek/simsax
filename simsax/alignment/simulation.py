import datetime
import itertools
import string
import gc

import numpy as np
import pandas as pd

import warnings
with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    from mpl_toolkits import mplot3d
    import matplotlib.gridspec as gridspec
    from matplotlib import colors as clrs

from simsax.alignment.calculators import AlignmentsCoverageCalculator
from simsax.alignment.finders import calculate_global_project_alignment, create_sax_signature, find_motifs, calculate_project_alignment
from simsax.alignment.plot import plot_all_motif_alignments, plot_project_sequence_with_motifs
from sax.pysax import SAXModel

from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score

from openpyxl import Workbook


def fill_motif_entries_in_xlsx(wb, motifs, project_sequence, title,
                               window, nbins, alphabet, score_threshold, within_min_dist,
                               stride):
    motifs_ws = wb.create_sheet(title)
    motifs_ws.cell(row=1, column=1).value = "motif"
    motifs_ws.cell(row=1, column=2).value = "start_idx"
    motifs_ws.cell(row=1, column=3).value = "end_idx"

    motifs_ws.cell(row=1, column=4).value = "project"
    motifs_ws.cell(row=1, column=5).value = "window"
    motifs_ws.cell(row=1, column=6).value = "nbins"
    motifs_ws.cell(row=1, column=7).value = "alphabet"
    motifs_ws.cell(row=1, column=8).value = "score_threshold"
    motifs_ws.cell(row=1, column=9).value = "within_min_dist"
    motifs_ws.cell(row=1, column=10).value = "stride"


    row_index = 2
    for motif in motifs:
        for align in list(motif.alignments):
            if align.window_a.sequence.equals(project_sequence.sequence):
                motif_window = align.window_a
            else:
                motif_window = align.window_b

            motifs_ws.cell(row=row_index, column=1).value = str(motif)
            motifs_ws.cell(row=row_index, column=2).value = motif_window.start_index
            motifs_ws.cell(row=row_index, column=3).value = motif_window.end_index
            motifs_ws.cell(row=row_index, column=4).value = project_sequence.project.name
            motifs_ws.cell(row=row_index, column=5).value = window
            motifs_ws.cell(row=row_index, column=6).value = nbins
            motifs_ws.cell(row=row_index, column=7).value = alphabet
            motifs_ws.cell(row=row_index, column=8).value = score_threshold
            motifs_ws.cell(row=row_index, column=9).value = within_min_dist
            motifs_ws.cell(row=row_index, column=10).value = stride

            row_index += 1


    return motifs_ws



def save_comparison_results(file_path, window, nbins, alphabet, score_threshold, within_min_dist,
                            stride, calc, motifs_between,
                            project_a_seq, project_b_seq, perform_within_simulation=False):

    between_cov_a = calc.metrics.get(project_a_seq.project, {}).get('between_coverage', 0)
    between_cov_b = calc.metrics.get(project_b_seq.project, {}).get('between_coverage', 0)
    between_cov_avg = (between_cov_a + between_cov_b) / 2
    between_cov_50_50 = 1 if between_cov_a >= 0.50 and between_cov_b >= 0.50 else 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ws.cell(row=1, column=1).value = "project_a"
    ws.cell(row=1, column=2).value = "project_b"
    ws.cell(row=1, column=3).value = "window"
    ws.cell(row=1, column=4).value = "nbins"
    ws.cell(row=1, column=5).value = "alphabet"
    ws.cell(row=1, column=6).value = "score_threshold"
    ws.cell(row=1, column=7).value = "within_min_dist"
    ws.cell(row=1, column=8).value = "stride"
    ws.cell(row=1, column=9).value = "coverage_a"
    ws.cell(row=1, column=10).value = "coverage_b"
    ws.cell(row=1, column=11).value = "coverage_avg"
    ws.cell(row=1, column=12).value = "50_50_criterion"

    ws.cell(row=2, column=1).value = project_a_seq.project.name
    ws.cell(row=2, column=2).value = project_b_seq.project.name
    ws.cell(row=2, column=3).value = window
    ws.cell(row=2, column=4).value = nbins
    ws.cell(row=2, column=5).value = alphabet
    ws.cell(row=2, column=6).value = score_threshold
    ws.cell(row=2, column=7).value = within_min_dist
    ws.cell(row=2, column=8).value = stride
    ws.cell(row=2, column=9).value = between_cov_a
    ws.cell(row=2, column=10).value = between_cov_b
    ws.cell(row=2, column=11).value = between_cov_avg
    ws.cell(row=2, column=12).value = between_cov_50_50

    if perform_within_simulation == True:
        ws.cell(row=1, column=13).value = "within_coverage_a"
        ws.cell(row=1, column=14).value = "within_coverage_b"
        ws.cell(row=2, column=13).value = calc.metrics.get(project_a_seq.project, {}).get('within_coverage', 0)
        ws.cell(row=2, column=14).value = calc.metrics.get(project_b_seq.project, {}).get('within_coverage', 0)


    motif_a_between_ws = fill_motif_entries_in_xlsx(wb, motifs_between, project_a_seq, "Motif A Between",
                                                    window, nbins, alphabet, score_threshold, within_min_dist,
                                                    stride)
    motif_b_between_ws = fill_motif_entries_in_xlsx(wb, motifs_between, project_b_seq, "Motif B Between",
                                                    window, nbins, alphabet, score_threshold, within_min_dist,
                                                    stride)

    common_weeks_ws = wb.create_sheet("Between_Weeks")
    common_weeks_ws.cell(row=1, column=1).value = "project_a_weeks"
    common_weeks_ws.cell(row=1, column=2).value = "project_b_weeks"

    for row_index, week in enumerate(calc.between_covered_set.get(project_a_seq.project, []), 2):
        common_weeks_ws.cell(row=row_index, column=1).value = week
    for row_index, week in enumerate(calc.between_covered_set.get(project_b_seq.project, []), 2):
        common_weeks_ws.cell(row=row_index, column=2).value = week


    wb.save(file_path + ".xlsx")


def simulate_alignments(params_combinations, project_sequences, output_folder="./simulation/",
                        perform_within_simulation=False, check_50_50_criterion=True, ignore_list=None):


    #create an output xlsx file
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.cell(row=1, column=1).value = "project_a"
    ws.cell(row=1, column=2).value = "project_b"
    ws.cell(row=1, column=3).value = "window"
    ws.cell(row=1, column=4).value = "nbins"
    ws.cell(row=1, column=5).value = "alphabet"
    ws.cell(row=1, column=6).value = "score_threshold"
    ws.cell(row=1, column=7).value = "within_min_dist"
    ws.cell(row=1, column=8).value = "stride"
    ws.cell(row=1, column=9).value = "coverage_a"
    ws.cell(row=1, column=10).value = "coverage_b"
    ws.cell(row=1, column=11).value = "coverage_avg"
    ws.cell(row=1, column=12).value = "50_50_criterion"
    if perform_within_simulation == True:
        ws.cell(row=1, column=13).value = "within_coverage_a"
        ws.cell(row=1, column=14).value = "within_coverage_b"
    if ignore_list is not None:
        ws.cell(row=1, column=15).value = "accuracy"
        ws.cell(row=1, column=16).value = "fscore"
        ws.cell(row=1, column=17).value = "precision"
        ws.cell(row=1, column=18).value = "recall"


    t = datetime.datetime.now()
    xlsx_file_name = "sim_{}_{:02d}_{}_{}{}{}.xlsx".format(t.year, t.month, t.day, t.hour, t.minute, t.second)

    row_index = 2
    for params in params_combinations:
        gc.collect()

        # get parameters for the combination to analyze
        window, within_min_dist, nbins, alphabet, score_threshold, stride = params
        simulation_fingerprint = "win-{}_nbins-{}_alph-{}_score-{}_min-dist-{}_stride-{}".format(
            window, nbins, alphabet, score_threshold, within_min_dist, stride)

        within_min_dist_window = int(window * within_min_dist)

        if nbins > window:
            continue


        # Create a sax parsers based on the parameters
        sax = SAXModel(
            window=window,
            stride=stride,
            nbins=nbins,
            alphabet=string.ascii_lowercase[:alphabet],
            epsilon=0.05)

        # Iterate over combination of projects
        projects_pairs = list(itertools.combinations(project_sequences, 2))
        for project_pair in projects_pairs:
            project_a_seq, project_b_seq = project_pair

            if min(max(project_a_seq.sequence), max(project_b_seq.sequence)) < alphabet:
                continue

            # calculate windows for a given sax signature
            project_a_seq.transform_project_dna(sax)
            project_b_seq.transform_project_dna(sax)

            # find alignments between and within the projects for given set of params
            #alignments_between = calculate_global_project_alignment(
            alignments_between = calculate_project_alignment(
                project_a_seq, project_b_seq,
                create_sax_signature(sax), within_min_dist_window, score_threshold)

            if perform_within_simulation == True:

                #project_a_within = calculate_global_project_alignment(project_a_seq,
                project_a_within = calculate_project_alignment(project_a_seq,
                                                                  project_a_seq,
                                                                  create_sax_signature(
                                                                    sax),
                                                                  within_min_dist_window,
                                                                  score_threshold)
                #project_b_within = calculate_global_project_alignment(project_b_seq,
                project_b_within = calculate_project_alignment(project_b_seq,
                                                                  project_b_seq,
                                                                  create_sax_signature(
                                                                    sax),
                                                                  within_min_dist_window,
                                                                  score_threshold)
            else:
                project_a_within = []
                project_b_within = []

            # calculate metrics
            calc = AlignmentsCoverageCalculator(
                alignments_between,
                {project_a_seq.project: project_a_within,
                 project_b_seq.project: project_b_within})
            calc.calculate_weeks()

            if ignore_list is not None:
                actual = [0 if x in ignore_list else 1 for x in range(0, len(project_a_seq.sequence))] 
                if project_a_seq.project in calc.between_covered_set.keys():
                    covered = calc.between_covered_set[project_a_seq.project]
                    predicted = [1 if x in covered else 0 for x in range(0, len(project_a_seq.sequence))] 
                else:
                    predicted = [0 for x in range(0, len(project_a_seq.sequence))] 
                accuracy = accuracy_score(actual, predicted)
                fscore = f1_score(actual, predicted)
                precision = precision_score(actual, predicted)
                recall = recall_score(actual, predicted)
   

            between_cov_a = calc.metrics.get(project_a_seq.project, {}).get('between_coverage', 0)
            between_cov_b = calc.metrics.get(project_b_seq.project, {}).get('between_coverage', 0)
            between_cov_avg = (between_cov_a + between_cov_b) / 2
            between_cov_50_50 = 1 if between_cov_a >= 0.50 and between_cov_b >= 0.50 else 0

            ws.cell(row=row_index, column=1).value = project_a_seq.project.name[0:20]
            ws.cell(row=row_index, column=2).value = project_b_seq.project.name[0:20]
            ws.cell(row=row_index, column=3).value = window
            ws.cell(row=row_index, column=4).value = nbins
            ws.cell(row=row_index, column=5).value = alphabet
            ws.cell(row=row_index, column=6).value = score_threshold
            ws.cell(row=row_index, column=7).value = within_min_dist
            ws.cell(row=row_index, column=8).value = stride
            ws.cell(row=row_index, column=9).value = between_cov_a
            ws.cell(row=row_index, column=10).value = between_cov_b
            ws.cell(row=row_index, column=11).value = between_cov_avg
            ws.cell(row=row_index, column=12).value = between_cov_50_50
            if perform_within_simulation == True:
                ws.cell(row=row_index, column=13).value = calc.metrics.get(project_a_seq.project, {}).get('within_coverage', 0)
                ws.cell(row=row_index, column=14).value = calc.metrics.get(project_b_seq.project, {}).get('within_coverage', 0)
            if ignore_list is not None:
                ws.cell(row=row_index, column=15).value = accuracy
                ws.cell(row=row_index, column=16).value = fscore
                ws.cell(row=row_index, column=17).value = precision
                ws.cell(row=row_index, column=18).value = recall
            wb.save(output_folder + xlsx_file_name)
            print(window, nbins, alphabet, score_threshold, between_cov_a, between_cov_b)
            row_index += 1

            if check_50_50_criterion == False or between_cov_50_50:

                # find motifs
                motifs_between = find_motifs(alignments_between)
                if perform_within_simulation == True:
                    motifs_project_a_within = find_motifs(project_a_within)
                    motifs_project_b_within = find_motifs(project_b_within)
                else:
                    motifs_project_a_within = []
                    motifs_project_b_within = []

                file_path = "{}{}_{}_{}".format(output_folder, simulation_fingerprint,
                                                    project_a_seq.project.name[0:20], project_b_seq.project.name[0:20])

                save_comparison_results(file_path, window, nbins, alphabet, score_threshold, within_min_dist,
                            stride, calc, motifs_between,
                            project_a_seq, project_b_seq, perform_within_simulation)



                with PdfPages(file_path+".pdf") as pdf:
                    try:
                        project_a_title = "{} {}".format(project_a_seq.project.name[0:10], simulation_fingerprint)
                        fig = plot_project_sequence_with_motifs(project_a_seq,
                                                                                          motifs_within=motifs_project_a_within,
                                                                                          motifs_between=motifs_between,
                                                                                          axis_x_dates=False,
                                                                                          figsize=(15, 8),
                                                                                          ylabel="#changes",
                                                                                          title=project_a_title,
                                                                                          between_coverage="{:.2f}".format(
                                                                                              calc.metrics.get(
                                                                                                  project_a_seq.project,
                                                                                                  {'between_coverage':0})[
                                                                                                  'between_coverage']))

                        plt.tight_layout()
                        pdf.savefig()
                        plt.close()
                        project_b_title = "{} {}".format(project_b_seq.project.name[0:10], simulation_fingerprint)
                        fig = plot_project_sequence_with_motifs(project_b_seq,
                                                                                          motifs_within=motifs_project_b_within,
                                                                                          motifs_between=motifs_between,
                                                                                          axis_x_dates=False,
                                                                                          figsize=(15, 8),
                                                                                          ylabel="#changes",
                                                                                          title=project_b_title,
                                                                                          between_coverage="{:.2f}".format(
                                                                                              calc.metrics.get(
                                                                                                  project_b_seq.project,
                                                                                                  {'between_coverage':0})[
                                                                                                  'between_coverage']))
                        plt.tight_layout()
                        pdf.savefig()
                        plt.close()

                        if len(motifs_between) > 0:
                            plot_all_motif_alignments(motifs_within=[], motifs_between=motifs_between,
                                                      width=5, height=3)
                            plt.tight_layout()
                            pdf.savefig()
                            plt.close()
                    except BaseException as e:
                        print('An exception occurred: {}'.format(e))
    return xlsx_file_name

def plot_simulation_3d(simulation_data):
    fig = plt.figure(figsize=(12,16))

    gs = gridspec.GridSpec(3, 2, height_ratios=[2, 1, 1])
    gs.update(hspace=0.25)

    ax = plt.subplot(gs[0, :], projection='3d')

    zdata = simulation_data['fscore']
    xdata = simulation_data['nbins']
    ydata = simulation_data['window']
    cdata = simulation_data['alphabet'].astype('int')

    cmap = plt.cm.Blues(np.linspace(0,1,50))
    cmap = clrs.ListedColormap(cmap[20:40,:-1])

    ax.scatter3D(xdata, ydata, zdata, c=cdata, cmap=cmap, depthshade=True);

    ax.set_xlabel('word length (w)')
    ax.set_ylabel('window length (n)')
    ax.set_zlabel('F-score')

    ax.set_xticks(np.arange(0, 66, 4))
    ax.set_yticks([4, 8, 16, 32, 64], minor=False)
    ax.set_yticks(np.arange(4,64,1), minor=True)

    ax.set_xlim(4, 64)

    fig.colorbar(ax.get_children()[1], label='alphabet size (a)', 
                 pad=0.0, orientation="vertical",
                 fraction=0.02, ticks=np.arange(4,21,1))

    boxprops = dict(linestyle='-', linewidth=1, color='k')
    medianprops = dict(linestyle='-', linewidth=4, color='black')
    flierprops = dict(marker='o', markerfacecolor='white', markersize=2,
                      linestyle='none')

    ax = plt.subplot(gs[1,0])
    bp = simulation_data.boxplot(ax=ax, column=['fscore'], by=['window'], showfliers=True,
                                boxprops=boxprops,
                                medianprops=medianprops, flierprops=flierprops, patch_artist=True)
    plt.suptitle('')
    ax.set_title(r'F-score by window length ($n$)')
    ax.set_xlabel('window length (n)')
    ax.set_ylabel('F-score')


    ax = plt.subplot(gs[1,1])
    simulation_data.boxplot(ax=ax, column=['fscore'], by=['alphabet'], showfliers=True,
                           boxprops=boxprops,
                                medianprops=medianprops, flierprops=flierprops, patch_artist=True)
    plt.suptitle('')
    ax.set_title(r'F-score by alphabet size ($a$)')
    ax.set_xlabel('alphabet size (a)')
    ax.set_ylabel('F-score')

    ax = plt.subplot(gs[2, :])
    simulation_data.boxplot(ax=ax, column=['fscore'], by=['nbins'], showfliers=True,
                           boxprops=boxprops,
                                medianprops=medianprops, flierprops=flierprops, patch_artist=True)
    plt.suptitle('')
    ax.set_title(r'F-score by word length ($w$)')
    ax.set_xlabel('word length (w)')
    ax.set_ylabel('F-score')

    plt.show()
    plt.close()

def plot_simulation_word_vs_w_to_n(simulation_data):
    fig = plt.figure(figsize=(15,5))

    zdata = simulation_data['fscore']
    xdata = simulation_data['nbins']
    ydata = simulation_data['window']
    cdata = simulation_data['alphabet'].astype('int')

    ax = plt.subplot(111)

    cmap = plt.cm.Reds(np.linspace(0,1,50))
    cmap = clrs.ListedColormap(cmap[20:40,:-1])

    max_fscore = simulation_data['fscore'].astype('float32').max()

    for alphabet in range(3, 21):
        data = simulation_data[(simulation_data['fscore'].astype('float32') == max_fscore) & (simulation_data['alphabet'] == alphabet)]

        xdata =  data.groupby('window').min().index.tolist()
        ydata = data.groupby('window').min()['nbins'] / data.groupby('window').min().index.tolist()
        cdata = data['fscore']
        ax.plot(xdata, ydata, c=cmap(alphabet), alpha=1, marker='o', label=f'{alphabet}')

        _ = ax.set_xticks(simulation_data['window'].unique().tolist())
        _ = ax.set_yticks(np.arange(0,1.01,0.1))

        ax.set_title(f'Minimum w/n ratio for F-score = {max_fscore:.2f}')
        ax.set_xlabel('window length (n)')
        ax.set_ylabel('min(w/n)')
        ax.grid(True)

    plt.legend(loc='upper right', frameon=True, ncol=4, 
               fancybox=True, shadow=False, framealpha=1, borderpad=0.5, edgecolor='gray',
               fontsize='x-small', title='alphabet', labelspacing=0.5, columnspacing=0.5)

    plt.show()
    plt.close()
    
